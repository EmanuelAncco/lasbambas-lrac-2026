"""
ANALISIS COMPLETO LRAC - CASO MMG / MINA JUANITA S.A.
Para informe Innovadores en Accion - Caso 13 VP SHE - Diagnostico y Gobernanza LRAC
Autor: Emanuel Edgar Ancco Guaygua
Fecha: 2026-05-20
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from statistics import mean, stdev, median
from collections import defaultdict
import json

EXCEL_PATH = r"C:/Users/Emanuel/Downloads/00 Data (10).xlsx"
OUT_DIR = r"C:/Users/Emanuel/.gemini/antigravity/scratch/n8n/lasbambas-lrac-2026/analysis"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ===================================================================
# CARGAR DATOS
# ===================================================================
base_rows = []
for row in wb['BASE'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        base_rows.append({
            'VP': row[0], 'GER': row[1], 'PILAR': row[2],
            'HERR': row[3], 'VALOR': row[4], 'MES': row[5],
            'ANIO': row[6], 'MESNUM': row[7]
        })

vp_rows = []
for row in wb['LRACxVP'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        vp_rows.append({
            'VP': row[0], 'MES': row[2], 'MESNUM': row[3],
            'FTO': row[5], '3Q': row[6], 'CCV': row[7], 'ACC': row[8],
            'NMAP': row[9], 'VEA': row[10], 'DESEMP': row[11],
            'L': row[12], 'R': row[13], 'A': row[14], 'C': row[15],
            'LRAC4P': row[16]
        })

ger_rows = []
for row in wb['LRACxGerencia'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        ger_rows.append({
            'VP': row[0], 'GER': row[1], 'MES': row[3], 'MESNUM': row[4],
            '3Q': row[6], 'ACC': row[7], 'CCV': row[8], 'DESEMP': row[9],
            'FTO': row[10], 'NMAP': row[11], 'VEA': row[12],
            'L': row[13], 'R': row[14], 'A': row[15], 'C': row[16],
            'LRAC4P': row[17]
        })

print(f"Filas BASE: {len(base_rows)}")
print(f"Filas LRACxVP: {len(vp_rows)}")
print(f"Filas LRACxGerencia: {len(ger_rows)}")

# Estructura organizacional
struct = defaultdict(set)
for r in base_rows:
    struct[r['VP']].add(r['GER'])

print("\n=== ESTRUCTURA ORGANIZACIONAL MINA JUANITA S.A. ===")
total_gerencias = 0
for vp in sorted(struct.keys()):
    gs = sorted(struct[vp])
    print(f"VP {vp}: {len(gs)} gerencias")
    for g in gs:
        print(f"   - {g}")
    total_gerencias += len(gs)
print(f"Total gerencias: {total_gerencias}")

# ===================================================================
# ANALISIS 1: ENERO 2026 - PEOR GERENCIA POR HERRAMIENTA
# ===================================================================
print("\n" + "="*80)
print("ANALISIS 1: ENERO 2026 - PEOR GERENCIA POR HERRAMIENTA")
print("="*80)

resultados_a1 = {}
for h in ['FTO', '3Q', 'ACC', 'NMAP']:
    enero_h = [r for r in base_rows if r['MES']=='ENERO' and r['HERR']==h and r['VALOR'] is not None]
    enero_h.sort(key=lambda r: r['VALOR'])
    peor = enero_h[0]
    resultados_a1[h] = {
        'peor_gerencia': peor['GER'],
        'peor_vp': peor['VP'],
        'peor_valor': peor['VALOR'],
        'ranking_top3_peores': [{'vp': r['VP'], 'ger': r['GER'], 'valor': r['VALOR']} for r in enero_h[:3]]
    }
    print(f"\n{h}: PEOR = {peor['GER']} (VP {peor['VP']}) con {peor['VALOR']*100:.1f}%")
    print(f"  Top 3 peores en {h}:")
    for r in enero_h[:5]:
        print(f"    {r['VP']:12s} | {r['GER']:25s} | {r['VALOR']*100:5.1f}%")
    print(f"  Media nacional {h} enero: {mean([r['VALOR'] for r in enero_h])*100:5.2f}%")
    print(f"  Mediana nacional {h} enero: {median([r['VALOR'] for r in enero_h])*100:5.2f}%")
    print(f"  Desv std nacional {h} enero: {stdev([r['VALOR'] for r in enero_h])*100:5.2f} pp")

# ===================================================================
# ANALISIS 2: ANIO 2026 GENERAL - VP CON PROMEDIO MAS ALTO
# ===================================================================
print("\n" + "="*80)
print("ANALISIS 2: AÑO 2026 GENERAL - VP CON LRAC-4P MAS ALTO")
print("="*80)

vps = sorted(set(r['VP'] for r in vp_rows))
vp_summary = []
for vp in vps:
    rows = [r for r in vp_rows if r['VP'] == vp]
    lrac_vals = [r['LRAC4P'] for r in rows]
    prom = mean(lrac_vals)
    desv = stdev(lrac_vals) if len(lrac_vals) > 1 else 0
    cv = desv/prom * 100 if prom else 0
    vp_summary.append({
        'vp': vp, 'promedio': prom, 'desv': desv, 'cv': cv,
        'min': min(lrac_vals), 'max': max(lrac_vals),
        'amplitud': max(lrac_vals) - min(lrac_vals),
        'detalle_mensual': {r['MES']: r['LRAC4P'] for r in sorted(rows, key=lambda x: x['MESNUM'])}
    })

vp_summary.sort(key=lambda x: x['promedio'], reverse=True)

print("\nRanking VPs por LRAC-4P anual:")
print(f"{'VP':<12} {'µ_anual':>10} {'σ':>8} {'CV%':>8} {'min':>8} {'max':>8} {'amp':>8}")
for v in vp_summary:
    print(f"{v['vp']:<12} {v['promedio']*100:>9.2f}% {v['desv']*100:>7.2f} {v['cv']:>7.2f} {v['min']*100:>7.2f}% {v['max']*100:>7.2f}% {v['amplitud']*100:>7.2f}")

vp_top = vp_summary[0]
print(f"\n>>> VP con promedio LRAC-4P MAS ALTO: {vp_top['vp']} con {vp_top['promedio']*100:.2f}%")
print(f"    Estabilidad: CV={vp_top['cv']:.2f}%, σ={vp_top['desv']*100:.2f}pp, amplitud={vp_top['amplitud']*100:.2f}pp")

# Indicador mas alto dentro de VP top
rows_top = [r for r in vp_rows if r['VP'] == vp_top['vp']]
indicadores = ['FTO', '3Q', 'CCV', 'ACC', 'NMAP', 'VEA', 'DESEMP']
ind_summary = []
for ind in indicadores:
    vals = [r[ind] for r in rows_top]
    ind_summary.append({'ind': ind, 'prom': mean(vals), 'desv': stdev(vals) if len(vals)>1 else 0, 'min': min(vals), 'max': max(vals)})

print(f"\nDesglose herramientas en VP {vp_top['vp']}:")
print(f"{'Herr':<10} {'µ':>10} {'σ':>8} {'CV%':>8} {'min':>8} {'max':>8}")
for i in sorted(ind_summary, key=lambda x: x['prom'], reverse=True):
    cv_i = i['desv']/i['prom']*100 if i['prom'] else 0
    print(f"{i['ind']:<10} {i['prom']*100:>9.2f}% {i['desv']*100:>7.2f} {cv_i:>7.2f} {i['min']*100:>7.2f}% {i['max']*100:>7.2f}%")

top_ind = max(ind_summary, key=lambda x: x['prom'])
print(f"\n>>> Indicador MAS ALTO en VP {vp_top['vp']}: {top_ind['ind']} con {top_ind['prom']*100:.2f}%")

# ===================================================================
# ANALISIS 3: PROYECTOS - PROMEDIO + ACC
# ===================================================================
print("\n" + "="*80)
print("ANALISIS 3: PROYECTOS AÑO 2026 - PROMEDIO GENERAL + ACC")
print("="*80)

rows_proj = [r for r in vp_rows if r['VP'] == 'PROYECTOS']
print(f"\nFilas PROYECTOS: {len(rows_proj)} meses (Ene-Abr 2026)")

# LRAC-4P PROYECTOS
lrac_proj = [r['LRAC4P'] for r in rows_proj]
print(f"\nLRAC-4P PROYECTOS:")
print(f"  Promedio anual: {mean(lrac_proj)*100:.2f}%")
print(f"  Detalle mensual:")
for r in sorted(rows_proj, key=lambda x: x['MESNUM']):
    print(f"    {r['MES']:10s} = {r['LRAC4P']*100:5.2f}%")

# Promedio de cada herramienta en PROYECTOS
print(f"\nIndicadores PROYECTOS 2026 (promedio Ene-Abr):")
proj_ind = []
for ind in indicadores:
    vals = [r[ind] for r in rows_proj]
    proj_ind.append({'ind': ind, 'prom': mean(vals), 'desv': stdev(vals) if len(vals)>1 else 0})
print(f"{'Herr':<10} {'µ':>10} {'σ':>8} {'min':>8} {'max':>8}")
for i in sorted(proj_ind, key=lambda x: x['prom'], reverse=True):
    vals = [r[i['ind']] for r in rows_proj]
    print(f"{i['ind']:<10} {i['prom']*100:>9.2f}% {i['desv']*100:>7.2f} {min(vals)*100:>7.2f}% {max(vals)*100:>7.2f}%")

# ACC especifico PROYECTOS
acc_proj = [r['ACC'] for r in rows_proj]
print(f"\n>>> ACC en PROYECTOS:")
print(f"    Promedio anual: {mean(acc_proj)*100:.2f}%")
print(f"    Detalle:")
for r in sorted(rows_proj, key=lambda x: x['MESNUM']):
    print(f"      {r['MES']:10s} = {r['ACC']*100:5.2f}%")
print(f"    Mes peor: {min(rows_proj, key=lambda x: x['ACC'])['MES']} ({min(acc_proj)*100:.2f}%)")
print(f"    Mes mejor: {max(rows_proj, key=lambda x: x['ACC'])['MES']} ({max(acc_proj)*100:.2f}%)")

# Promedio general PROYECTOS (todos los indicadores)
all_proj = []
for r in rows_proj:
    for ind in indicadores:
        all_proj.append(r[ind])
print(f"\n>>> Promedio general PROYECTOS (todos indicadores x meses, n={len(all_proj)}): {mean(all_proj)*100:.2f}%")

# Pilares PROYECTOS
print(f"\nPilares PROYECTOS 2026:")
for pil in ['L', 'R', 'A', 'C']:
    vals = [r[pil] for r in rows_proj]
    print(f"  Pilar {pil}: µ={mean(vals)*100:5.2f}%  rango=[{min(vals)*100:.2f}%, {max(vals)*100:.2f}%]")

# ===================================================================
# ANALISIS 4: INDICE LRAC POR VP - ALTO Y BAJO AL CIERRE
# ===================================================================
print("\n" + "="*80)
print("ANALISIS 4: INDICE LRAC POR VP - PROMEDIO ANUAL Y CIERRE (ABRIL)")
print("="*80)

# Promedio anual ya calculado en vp_summary
print(f"\nA) Por PROMEDIO ANUAL Ene-Abr 2026:")
print(f"   MAS ALTO: {vp_summary[0]['vp']} con {vp_summary[0]['promedio']*100:.2f}%")
print(f"   MAS BAJO: {vp_summary[-1]['vp']} con {vp_summary[-1]['promedio']*100:.2f}%")
print(f"   Brecha: {(vp_summary[0]['promedio']-vp_summary[-1]['promedio'])*100:.2f} puntos porcentuales")

# Al cierre = ABRIL
abril = [r for r in vp_rows if r['MES']=='ABRIL']
abril_sorted = sorted(abril, key=lambda r: r['LRAC4P'], reverse=True)
print(f"\nB) Al CIERRE del periodo (ABRIL 2026):")
for r in abril_sorted:
    print(f"   {r['VP']:<12}: LRAC-4P = {r['LRAC4P']*100:.2f}%")
print(f"   MAS ALTO: {abril_sorted[0]['VP']} con {abril_sorted[0]['LRAC4P']*100:.2f}%")
print(f"   MAS BAJO: {abril_sorted[-1]['VP']} con {abril_sorted[-1]['LRAC4P']*100:.2f}%")

# Evolucion mensual cada VP
print(f"\nC) Evolucion mensual LRAC-4P por VP:")
print(f"{'VP':<12} {'ENERO':>10} {'FEB':>10} {'MAR':>10} {'ABR':>10} {'Δ Ene→Abr':>12}")
for vp in sorted(vps):
    rows_vp = sorted([r for r in vp_rows if r['VP']==vp], key=lambda x: x['MESNUM'])
    ene_v = rows_vp[0]['LRAC4P']
    abr_v = rows_vp[-1]['LRAC4P']
    delta = (abr_v - ene_v) * 100
    print(f"{vp:<12} {ene_v*100:>9.2f}% {rows_vp[1]['LRAC4P']*100:>9.2f}% {rows_vp[2]['LRAC4P']*100:>9.2f}% {abr_v*100:>9.2f}% {delta:>+11.2f}pp")

# ===================================================================
# ANALISIS 5: DESCUBRIMIENTOS ADICIONALES (para enriquecer informe)
# ===================================================================
print("\n" + "="*80)
print("ANALISIS 5: HALLAZGOS ADICIONALES PARA INFORME")
print("="*80)

# Estado por gerencia-mes (clasificacion)
print("\n5.1 - Distribucion de estados LRAC-4P por gerencia-mes:")
states_count = defaultdict(int)
for r in ger_rows:
    # Clasificacion: usar bandas del codigo COLOR del Excel (que ya esta como Optimo - todos),
    # pero hagamos bandas conceptuales
    v = r['LRAC4P']
    if v >= 0.75: state = 'Óptimo (≥75%)'
    elif v >= 0.6: state = 'Esperado (60-75%)'
    else: state = 'Debajo (<60%)'
    states_count[state] += 1
for st, cnt in states_count.items():
    print(f"  {st}: {cnt} gerencia-mes ({cnt/len(ger_rows)*100:.1f}%)")

# Gerencias criticas (con LRAC-4P consistentemente bajo)
print("\n5.2 - Gerencias mas debiles (promedio LRAC-4P anual):")
ger_summary = defaultdict(list)
for r in ger_rows:
    ger_summary[(r['VP'], r['GER'])].append(r['LRAC4P'])
ger_list = [(vp, g, mean(vs), stdev(vs) if len(vs)>1 else 0) for (vp,g), vs in ger_summary.items()]
ger_list.sort(key=lambda x: x[2])
print(f"{'VP':<12} {'GERENCIA':<25} {'µ':>9} {'σ':>8}")
for vp, g, m, s in ger_list[:5]:
    print(f"{vp:<12} {g:<25} {m*100:>8.2f}% {s*100:>7.2f}")
print("\n  Gerencias mas fuertes:")
for vp, g, m, s in ger_list[-5:]:
    print(f"{vp:<12} {g:<25} {m*100:>8.2f}% {s*100:>7.2f}")

# Pilares - heatmap por VP
print("\n5.3 - Heatmap pilares promedio anual por VP:")
print(f"{'VP':<12} {'Pilar L':>10} {'Pilar R':>10} {'Pilar A':>10} {'Pilar C':>10}")
for vp in sorted(vps):
    rows = [r for r in vp_rows if r['VP']==vp]
    pl = mean([r['L'] for r in rows])
    pr = mean([r['R'] for r in rows])
    pa = mean([r['A'] for r in rows])
    pc = mean([r['C'] for r in rows])
    print(f"{vp:<12} {pl*100:>9.2f}% {pr*100:>9.2f}% {pa*100:>9.2f}% {pc*100:>9.2f}%")

# Pilares con mayor variabilidad (oportunidad)
print("\n5.4 - Pilares con MAYOR variabilidad (interVP) = mayor oportunidad estandarizacion:")
pilares_var = []
for pil in ['L', 'R', 'A', 'C']:
    # Promedio por VP
    vals_vp = [mean([r[pil] for r in vp_rows if r['VP']==vp]) for vp in vps]
    pilares_var.append((pil, stdev(vals_vp), max(vals_vp)-min(vals_vp)))
pilares_var.sort(key=lambda x: x[1], reverse=True)
for p, s, amp in pilares_var:
    print(f"  Pilar {p}: σ_interVP = {s*100:.2f}pp, amplitud = {amp*100:.2f}pp")

# Correlacion meta vs realidad - asumimos que >= 0.70 = optimo
print("\n5.5 - Brecha vs meta corporativa (asumimos meta 70%):")
META = 0.70
for vp in sorted(vps):
    rows = [r for r in vp_rows if r['VP']==vp]
    brecha = mean([r['LRAC4P'] for r in rows]) - META
    print(f"  {vp:<12}: brecha = {brecha*100:+.2f}pp")

# Indicador mas debil global
print("\n5.6 - Indicador mas debil a nivel CORPORATIVO (promedio anual):")
ind_global = []
for ind in indicadores:
    vals = [r[ind] for r in vp_rows]
    ind_global.append((ind, mean(vals), stdev(vals)))
ind_global.sort(key=lambda x: x[1])
for ind, m, s in ind_global:
    print(f"  {ind:<10}: µ={m*100:.2f}%  σ={s*100:.2f}pp")

print("\n" + "="*80)
print("FIN ANALISIS")
print("="*80)

# Guardar JSON para uso posterior
output = {
    'estructura_organizacional': {vp: sorted(list(gs)) for vp, gs in struct.items()},
    'total_gerencias': total_gerencias,
    'analisis_1_enero': resultados_a1,
    'vp_summary_anual': vp_summary,
    'vp_top_indicador_mas_alto': {'vp': vp_top['vp'], 'ind': top_ind['ind'], 'valor': top_ind['prom']},
    'proyectos_lrac_promedio': mean(lrac_proj),
    'proyectos_acc_promedio': mean(acc_proj),
    'proyectos_acc_por_mes': {r['MES']: r['ACC'] for r in rows_proj},
    'proyectos_promedio_general_indicadores': mean(all_proj),
    'abril_ranking': [{'vp': r['VP'], 'lrac': r['LRAC4P']} for r in abril_sorted],
}
with open(OUT_DIR + '/resultados.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, default=str, ensure_ascii=False)
print(f"\nResultados guardados en: {OUT_DIR}/resultados.json")
