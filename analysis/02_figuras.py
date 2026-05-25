"""
Generacion de figuras para informe LRAC.
Salida en figures/ como PNG alta calidad.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from collections import defaultdict
from statistics import mean, stdev

EXCEL = r"C:/Users/Emanuel/Downloads/00 Data (10).xlsx"
OUT = r"C:/Users/Emanuel/.gemini/antigravity/scratch/n8n/lasbambas-lrac-2026/figures/"

plt.rcParams.update({
    'font.family': 'serif',
    'font.size': 10,
    'figure.dpi': 150,
    'savefig.dpi': 200,
    'savefig.bbox': 'tight',
})

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# ---------- CARGAR DATOS ----------
base_rows = [{'VP': r[0], 'GER': r[1], 'PILAR': r[2], 'HERR': r[3], 'VALOR': r[4],
              'MES': r[5], 'MESNUM': r[7]} for r in wb['BASE'].iter_rows(min_row=2, values_only=True) if r[0]]

vp_rows = []
for row in wb['LRACxVP'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        vp_rows.append({
            'VP': row[0], 'MES': row[2], 'MESNUM': row[3],
            'FTO': row[5], '3Q': row[6], 'CCV': row[7], 'ACC': row[8],
            'NMAP': row[9], 'VEA': row[10], 'DESEMP': row[11],
            'L': row[12], 'R': row[13], 'A': row[14], 'C': row[15],
            'LRAC4P': row[16]
        })

ger_rows = []
for row in wb['LRACxGerencia'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        ger_rows.append({
            'VP': row[0], 'GER': row[1], 'MES': row[3], 'MESNUM': row[4],
            '3Q': row[6], 'ACC': row[7], 'CCV': row[8], 'DESEMP': row[9],
            'FTO': row[10], 'NMAP': row[11], 'VEA': row[12],
            'L': row[13], 'R': row[14], 'A': row[15], 'C': row[16],
            'LRAC4P': row[17]
        })

vps = sorted(set(r['VP'] for r in vp_rows))
meses_ord = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL']

# Paleta MMG: rojo institucional + grises
COLORES_VP = {
    'OPERACIONES': '#C8102E',  # rojo MMG
    'PROYECTOS':   '#FF7F0E',  # naranja
    'SHE':         '#2E7D32',  # verde
    'SUPPLY':      '#1F77B4',  # azul
}
PILARES_COLOR = {'L': '#C8102E', 'R': '#FF7F0E', 'A': '#2E7D32', 'C': '#1F77B4'}

# ==========================================================
# FIG 1: Evolucion LRAC-4P por VP (lineas)
# ==========================================================
fig, ax = plt.subplots(figsize=(8, 4.5))
x = np.arange(4)
for vp in vps:
    vals = [next(r['LRAC4P'] for r in vp_rows if r['VP']==vp and r['MES']==m)*100 for m in meses_ord]
    ax.plot(x, vals, marker='o', linewidth=2, label=vp, color=COLORES_VP.get(vp, 'gray'))
    for i, v in enumerate(vals):
        ax.text(i, v+0.5, f'{v:.1f}%', ha='center', fontsize=7)
ax.axhline(70, color='#888', linestyle='--', alpha=0.6, label='Meta corporativa 70%')
ax.set_xticks(x); ax.set_xticklabels(meses_ord)
ax.set_ylabel('Indice LRAC-4P (%)')
ax.set_title('Figura 1. Evolucion mensual del indice LRAC-4P por Vicepresidencia (Mina Juanita S.A., Ene-Abr 2026)')
ax.legend(loc='lower right', fontsize=8)
ax.set_ylim(60, 80)
ax.grid(alpha=0.3)
plt.savefig(OUT + 'fig01_evolucion_vp.png')
plt.close()
print('Fig 1 OK')

# ==========================================================
# FIG 2: Heatmap pilares por VP (anual)
# ==========================================================
fig, ax = plt.subplots(figsize=(7, 3.5))
pilares = ['L', 'R', 'A', 'C']
matriz = np.zeros((len(vps), len(pilares)))
for i, vp in enumerate(vps):
    rows = [r for r in vp_rows if r['VP']==vp]
    for j, p in enumerate(pilares):
        matriz[i, j] = mean([r[p] for r in rows]) * 100

im = ax.imshow(matriz, aspect='auto', cmap='RdYlGn', vmin=60, vmax=80)
ax.set_xticks(range(len(pilares)))
ax.set_xticklabels(['L (Liderazgo)', 'R (Riesgos)', 'A (Aprendizaje)', 'C (Contratistas)'])
ax.set_yticks(range(len(vps)))
ax.set_yticklabels(vps)
for i in range(len(vps)):
    for j in range(len(pilares)):
        ax.text(j, i, f'{matriz[i,j]:.1f}%', ha='center', va='center',
                color='black', fontweight='bold', fontsize=10)
cbar = plt.colorbar(im, ax=ax, fraction=0.04)
cbar.set_label('%')
ax.set_title('Figura 2. Heatmap del cumplimiento promedio por Pilar y VP (anual 2026)')
plt.savefig(OUT + 'fig02_heatmap_pilares.png')
plt.close()
print('Fig 2 OK')

# ==========================================================
# FIG 3: Barras peor gerencia por herramienta (Enero)
# ==========================================================
fig, axes = plt.subplots(1, 4, figsize=(13, 4))
herr_list = ['FTO', '3Q', 'ACC', 'NMAP']
for ax, h in zip(axes, herr_list):
    enero_h = [r for r in base_rows if r['MES']=='ENERO' and r['HERR']==h and r['VALOR'] is not None]
    enero_h.sort(key=lambda r: r['VALOR'])
    labels = [f"{r['GER'][:14]}\n({r['VP'][:5]})" for r in enero_h[:6]]
    vals = [r['VALOR']*100 for r in enero_h[:6]]
    colors = ['#C8102E' if i==0 else '#FF7F0E' if i<3 else '#999' for i in range(len(vals))]
    bars = ax.barh(range(len(vals)), vals, color=colors)
    ax.set_yticks(range(len(vals)))
    ax.set_yticklabels(labels, fontsize=7)
    ax.set_xlabel('% cumplimiento')
    ax.set_title(f'{h} - Enero 2026', fontsize=10, fontweight='bold')
    ax.invert_yaxis()
    for i, v in enumerate(vals):
        ax.text(v+0.5, i, f'{v:.1f}%', va='center', fontsize=7)
    ax.set_xlim(0, 100)
    ax.axvline(60, color='gray', linestyle=':', alpha=0.6)
plt.suptitle('Figura 3. Peores gerencias por herramienta en Enero 2026 (6 mas bajas)', fontsize=11)
plt.tight_layout()
plt.savefig(OUT + 'fig03_peor_gerencia_enero.png')
plt.close()
print('Fig 3 OK')

# ==========================================================
# FIG 4: Boxplot dispersion intra-VP
# ==========================================================
fig, ax = plt.subplots(figsize=(7, 4))
data = []
for vp in vps:
    vals = [r['LRAC4P']*100 for r in ger_rows if r['VP']==vp]
    data.append(vals)
bp = ax.boxplot(data, tick_labels=vps, patch_artist=True, widths=0.6,
                medianprops={'color':'black', 'linewidth':2})
for patch, vp in zip(bp['boxes'], vps):
    patch.set_facecolor(COLORES_VP.get(vp, 'gray'))
    patch.set_alpha(0.7)
ax.axhline(70, color='#888', linestyle='--', alpha=0.6)
ax.set_ylabel('LRAC-4P por gerencia-mes (%)')
ax.set_title('Figura 4. Dispersion del LRAC-4P intra-VP (gerencias × meses, 2026)')
ax.grid(alpha=0.3, axis='y')
plt.savefig(OUT + 'fig04_boxplot_dispersion.png')
plt.close()
print('Fig 4 OK')

# ==========================================================
# FIG 5: Radar 7 indicadores por VP
# ==========================================================
indicadores = ['FTO', '3Q', 'CCV', 'ACC', 'NMAP', 'VEA', 'DESEMP']
angles = np.linspace(0, 2*np.pi, len(indicadores), endpoint=False).tolist()
angles += angles[:1]
fig, ax = plt.subplots(figsize=(7, 7), subplot_kw=dict(projection='polar'))
for vp in vps:
    rows = [r for r in vp_rows if r['VP']==vp]
    vals = [mean([r[i] for r in rows])*100 for i in indicadores]
    vals += vals[:1]
    ax.plot(angles, vals, 'o-', linewidth=2, label=vp, color=COLORES_VP.get(vp, 'gray'))
    ax.fill(angles, vals, alpha=0.1, color=COLORES_VP.get(vp, 'gray'))
ax.set_theta_offset(np.pi/2)
ax.set_theta_direction(-1)
ax.set_thetagrids(np.degrees(angles[:-1]), indicadores)
ax.set_ylim(50, 85)
ax.set_yticks([60, 70, 80])
ax.set_yticklabels(['60%', '70%', '80%'])
ax.set_title('Figura 5. Perfil de 7 herramientas por VP (promedio anual 2026)', pad=20)
ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.05), fontsize=9)
plt.savefig(OUT + 'fig05_radar_indicadores.png')
plt.close()
print('Fig 5 OK')

# ==========================================================
# FIG 6: ACC PROYECTOS - serie temporal
# ==========================================================
fig, ax = plt.subplots(figsize=(7, 4))
proj = [r for r in vp_rows if r['VP']=='PROYECTOS']
acc_p = [next(r['ACC'] for r in proj if r['MES']==m)*100 for m in meses_ord]
lrac_p = [next(r['LRAC4P'] for r in proj if r['MES']==m)*100 for m in meses_ord]
x = np.arange(4)
ax.bar(x-0.2, acc_p, width=0.4, label='ACC PROYECTOS', color='#FF7F0E', alpha=0.8)
ax.bar(x+0.2, lrac_p, width=0.4, label='LRAC-4P PROYECTOS', color='#1F77B4', alpha=0.8)
for i, (a, l) in enumerate(zip(acc_p, lrac_p)):
    ax.text(i-0.2, a+1, f'{a:.1f}%', ha='center', fontsize=7)
    ax.text(i+0.2, l+1, f'{l:.1f}%', ha='center', fontsize=7)
ax.axhline(70, color='gray', linestyle='--', alpha=0.6, label='Meta 70%')
ax.set_xticks(x); ax.set_xticklabels(meses_ord)
ax.set_ylabel('%')
ax.set_title('Figura 6. ACC vs LRAC-4P en VP PROYECTOS (2026)')
ax.legend(fontsize=8)
ax.set_ylim(0, 90)
ax.grid(alpha=0.3, axis='y')
plt.savefig(OUT + 'fig06_acc_proyectos.png')
plt.close()
print('Fig 6 OK')

# ==========================================================
# FIG 7: Ranking gerencias (todas las 12 ordenadas)
# ==========================================================
fig, ax = plt.subplots(figsize=(9, 5.5))
ger_summary = defaultdict(list)
for r in ger_rows:
    ger_summary[(r['VP'], r['GER'])].append(r['LRAC4P'])
ger_list = [(vp, g, mean(vs)*100, stdev(vs)*100 if len(vs)>1 else 0) for (vp,g), vs in ger_summary.items()]
ger_list.sort(key=lambda x: x[2])
labels = [f"{g}\n({vp})" for vp, g, m, s in ger_list]
vals = [m for vp,g,m,s in ger_list]
errs = [s for vp,g,m,s in ger_list]
colors = [COLORES_VP.get(vp, 'gray') for vp, g, m, s in ger_list]
y = range(len(ger_list))
ax.barh(y, vals, xerr=errs, color=colors, alpha=0.85, edgecolor='black', linewidth=0.5,
        error_kw={'ecolor':'#333', 'capsize':3, 'linewidth':1})
ax.set_yticks(y); ax.set_yticklabels(labels, fontsize=8)
ax.set_xlabel('LRAC-4P promedio anual (%) ± σ')
ax.axvline(70, color='gray', linestyle='--', alpha=0.6, label='Meta 70%')
ax.axvline(60, color='red', linestyle=':', alpha=0.5, label='Umbral critico 60%')
for i, v in enumerate(vals):
    ax.text(v+1, i, f'{v:.1f}%', va='center', fontsize=7)
ax.set_title('Figura 7. Ranking anual LRAC-4P por gerencia (12 gerencias, 2026)')
ax.legend(loc='lower right', fontsize=8)
ax.set_xlim(50, 90)
ax.grid(alpha=0.3, axis='x')
plt.tight_layout()
plt.savefig(OUT + 'fig07_ranking_gerencias.png')
plt.close()
print('Fig 7 OK')

# ==========================================================
# FIG 8: Indicador mas debil corporativo (NMAP)
# ==========================================================
fig, ax = plt.subplots(figsize=(8, 4))
ind_list = ['FTO', '3Q', 'CCV', 'ACC', 'NMAP', 'VEA', 'DESEMP']
ind_corp = []
for ind in ind_list:
    vals = [r[ind] for r in vp_rows]
    ind_corp.append((ind, mean(vals)*100, stdev(vals)*100 if len(vals)>1 else 0))
ind_corp.sort(key=lambda x: x[1])
labels = [i[0] for i in ind_corp]
vals = [i[1] for i in ind_corp]
errs = [i[2] for i in ind_corp]
colors_b = ['#C8102E' if v < 67 else '#FF7F0E' if v < 70 else '#2E7D32' for v in vals]
ax.bar(labels, vals, yerr=errs, color=colors_b, alpha=0.8, edgecolor='black', linewidth=0.5,
       error_kw={'ecolor':'#333', 'capsize':3, 'linewidth':1})
for i, (v, e) in enumerate(zip(vals, errs)):
    ax.text(i, v+e+0.5, f'{v:.1f}%', ha='center', fontsize=8, fontweight='bold')
ax.axhline(70, color='gray', linestyle='--', alpha=0.6, label='Meta 70%')
ax.set_ylabel('Promedio corporativo (%) ± σ')
ax.set_title('Figura 8. Ranking corporativo de herramientas LRAC (2026) — NMAP es la mas debil')
ax.set_ylim(55, 80)
ax.legend(fontsize=8)
ax.grid(alpha=0.3, axis='y')
plt.savefig(OUT + 'fig08_indicador_debil.png')
plt.close()
print('Fig 8 OK')

# ==========================================================
# FIG 9: Pareto - gerencias y su contribucion a la brecha
# ==========================================================
fig, ax = plt.subplots(figsize=(9, 4.5))
META = 70
brechas = []
for (vp, g), vs in ger_summary.items():
    prom = mean(vs) * 100
    brecha = META - prom
    if brecha > 0:
        brechas.append((vp, g, prom, brecha))
brechas.sort(key=lambda x: x[3], reverse=True)
labels = [f"{g}" for vp,g,p,b in brechas]
vals_brecha = [b for vp,g,p,b in brechas]
colors = [COLORES_VP.get(vp, 'gray') for vp,g,p,b in brechas]
ax.bar(labels, vals_brecha, color=colors, alpha=0.85, edgecolor='black', linewidth=0.5)
acum = np.cumsum(vals_brecha)
acum_pct = acum / acum[-1] * 100
ax2 = ax.twinx()
ax2.plot(labels, acum_pct, 'o-', color='#333', linewidth=2)
ax2.set_ylabel('% acumulado')
ax2.axhline(80, color='red', linestyle=':', alpha=0.6)
ax.set_ylabel('Brecha vs meta 70% (pp)')
ax.set_xticklabels(labels, rotation=30, ha='right', fontsize=8)
ax.set_title('Figura 9. Pareto de brechas LRAC-4P por gerencia (vs meta 70%)')
plt.tight_layout()
plt.savefig(OUT + 'fig09_pareto_brechas.png')
plt.close()
print('Fig 9 OK')

# ==========================================================
# FIG 10: Hudson Safety Culture Maturity Ladder con posicion estimada
# ==========================================================
fig, ax = plt.subplots(figsize=(9, 4))
niveles = ['Patologica', 'Reactiva', 'Calculativa', 'Proactiva', 'Generativa']
rangos = [(0, 30), (30, 50), (50, 70), (70, 85), (85, 100)]
colores_h = ['#8B0000', '#C8102E', '#FF7F0E', '#FFD700', '#2E7D32']
for i, (lvl, (lo, hi), c) in enumerate(zip(niveles, rangos, colores_h)):
    ax.barh(0, hi-lo, left=lo, color=c, alpha=0.7, edgecolor='black')
    ax.text((lo+hi)/2, 0, lvl, ha='center', va='center', fontsize=10, fontweight='bold', color='white')

vp_anual = sorted([(vp, mean([r['LRAC4P'] for r in vp_rows if r['VP']==vp])*100) for vp in vps], key=lambda x: x[1])
for i, (vp, v) in enumerate(vp_anual):
    ax.plot(v, 0.4, 'v', markersize=15, color=COLORES_VP.get(vp, 'gray'), markeredgecolor='black')
    ax.text(v, 0.55 + i*0.08, f'{vp}\n{v:.1f}%', ha='center', fontsize=8, color=COLORES_VP.get(vp, 'gray'), fontweight='bold')
ax.set_xlim(0, 100)
ax.set_ylim(-0.5, 1.1)
ax.set_xticks(range(0, 101, 10))
ax.set_yticks([])
ax.set_xlabel('% cumplimiento del sistema LRAC')
ax.set_title('Figura 10. Diagnostico cultural de Mina Juanita S.A. sobre la escala de Hudson')
ax.grid(alpha=0.3, axis='x')
plt.tight_layout()
plt.savefig(OUT + 'fig10_hudson_diagnostico.png')
plt.close()
print('Fig 10 OK')

print('Todas las figuras generadas en', OUT)
