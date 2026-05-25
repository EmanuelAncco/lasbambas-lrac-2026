"""
ITER 2A - Estadistica inferencial sobre el caso LRAC Mina Juanita S.A.

Tests aplicados:
- Kruskal-Wallis para diferencias entre VPs (datos % no-normales, n=4 por VP).
- Post-hoc de Dunn con correccion Bonferroni (implementacion propia).
- Control chart Shewhart (I-Chart) sobre LRAC-4P corporativo mensual.
- Tendencia Mann-Kendall por VP (con scipy.stats.kendalltau como aproximacion).
- Score compuesto S = mu - k*sigma con k=1.
- Matriz de correlacion 7x7 entre herramientas.
"""
import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import numpy as np
from scipy import stats
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from itertools import combinations
from statistics import mean, stdev
from collections import defaultdict

EXCEL = r"C:/Users/Emanuel/Downloads/00 Data (10).xlsx"
OUT_DIR = r"C:/Users/Emanuel/.gemini/antigravity/scratch/n8n/lasbambas-lrac-2026/analysis"
FIG_DIR = r"C:/Users/Emanuel/.gemini/antigravity/scratch/n8n/lasbambas-lrac-2026/figures"

plt.rcParams.update({
    'font.family': 'serif',
    'font.size': 10,
    'figure.dpi': 150,
    'savefig.dpi': 200,
    'savefig.bbox': 'tight',
})

COLORES_VP = {
    'OPERACIONES': '#C8102E',
    'PROYECTOS':   '#FF7F0E',
    'SHE':         '#2E7D32',
    'SUPPLY':      '#1F77B4',
}

# ===================================================================
# CARGAR DATOS
# ===================================================================
wb = openpyxl.load_workbook(EXCEL, data_only=True)

vp_rows = []
for row in wb['LRACxVP'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        vp_rows.append({
            'VP': row[0], 'MES': row[2], 'MESNUM': row[3],
            'FTO': row[5], '3Q': row[6], 'CCV': row[7], 'ACC': row[8],
            'NMAP': row[9], 'VEA': row[10], 'DESEMP': row[11],
            'L': row[12], 'R': row[13], 'A': row[14], 'C': row[15],
            'LRAC4P': row[16]
        })

ger_rows = []
for row in wb['LRACxGerencia'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        ger_rows.append({
            'VP': row[0], 'GER': row[1], 'MES': row[3], 'MESNUM': row[4],
            '3Q': row[6], 'ACC': row[7], 'CCV': row[8], 'DESEMP': row[9],
            'FTO': row[10], 'NMAP': row[11], 'VEA': row[12],
            'L': row[13], 'R': row[14], 'A': row[15], 'C': row[16],
            'LRAC4P': row[17]
        })

global_rows = []
for row in wb['LRAC_Global'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        global_rows.append({
            'MES': row[1], 'MESNUM': row[2],
            'FTO': row[4], '3Q': row[5], 'CCV': row[6], 'ACC': row[7],
            'NMAP': row[8], 'VEA': row[9], 'DESEMP': row[10],
            'L': row[11], 'R': row[12], 'A': row[13], 'C': row[14],
            'LRAC4P': row[15]
        })

vps = sorted(set(r['VP'] for r in vp_rows))
meses = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL']

results = {}

# ===================================================================
# 1. KRUSKAL-WALLIS sobre LRAC-4P por VP
# ===================================================================
print("="*80)
print("1. KRUSKAL-WALLIS - LRAC-4P entre VPs")
print("="*80)

# Usamos los datos a nivel gerencia-mes (mas observaciones)
samples_vp = {vp: [r['LRAC4P'] for r in ger_rows if r['VP']==vp] for vp in vps}
print("\nMuestras por VP (gerencia x mes):")
for vp, vals in samples_vp.items():
    print(f"  {vp:<12}: n={len(vals)}, µ={np.mean(vals)*100:.2f}%, σ={np.std(vals, ddof=1)*100:.2f}pp")

H, p_kw = stats.kruskal(*samples_vp.values())
print(f"\nH_observado = {H:.4f}")
print(f"p-value     = {p_kw:.6f}")
print(f"Conclusion: ", end='')
if p_kw < 0.05:
    print(f"Rechazo H0 al 5%. Hay diferencias estadisticamente significativas entre las VPs.")
else:
    print(f"NO rechazo H0 al 5%. NO se demuestra diferencia significativa entre VPs (alpha=0.05).")

results['kruskal_wallis'] = {'H': H, 'p_value': p_kw, 'reject_h0_5pct': p_kw < 0.05}

# ===================================================================
# 2. POST-HOC DUNN con Bonferroni (implementacion propia)
# ===================================================================
print("\n" + "="*80)
print("2. POST-HOC DUNN (Bonferroni)")
print("="*80)

# Combinar todos los datos y rankear
all_vals = []
all_groups = []
for vp, vals in samples_vp.items():
    for v in vals:
        all_vals.append(v)
        all_groups.append(vp)
all_vals = np.array(all_vals)
all_groups = np.array(all_groups)
n_total = len(all_vals)

# Ranks (manejo de empates con average)
ranks = stats.rankdata(all_vals)

# Suma de rangos por grupo
rank_sum = {vp: ranks[all_groups==vp].sum() for vp in vps}
n_per_vp = {vp: (all_groups==vp).sum() for vp in vps}
mean_rank = {vp: rank_sum[vp]/n_per_vp[vp] for vp in vps}

print(f"\nRangos promedio por VP:")
for vp in vps:
    print(f"  {vp:<12}: mean_rank = {mean_rank[vp]:.2f}, n={n_per_vp[vp]}")

# Estadistico de Dunn entre cada par
# z = (R_i - R_j) / sqrt( (N(N+1)/12) * (1/n_i + 1/n_j) )
# corregido por empates si los hay
unique_vals, counts = np.unique(all_vals, return_counts=True)
ties = counts[counts > 1]
tie_correction = sum(t**3 - t for t in ties) / (n_total - 1) if n_total > 1 else 0
sigma2 = ((n_total*(n_total+1) - tie_correction)/12)

pairs = list(combinations(vps, 2))
dunn_results = []
print(f"\nComparaciones pareadas (z, p-bonferroni para 6 comparaciones):")
print(f"{'VP_i':<12} {'VP_j':<12} {'z':>8} {'p_raw':>10} {'p_bonf':>10} {'sig 5%':>8}")
n_pairs = len(pairs)
for vp_i, vp_j in pairs:
    diff = mean_rank[vp_i] - mean_rank[vp_j]
    se = np.sqrt(sigma2 * (1/n_per_vp[vp_i] + 1/n_per_vp[vp_j]))
    z = diff / se
    # two-sided p-value
    p_raw = 2 * (1 - stats.norm.cdf(abs(z)))
    p_bonf = min(p_raw * n_pairs, 1.0)
    sig = 'SI' if p_bonf < 0.05 else 'no'
    dunn_results.append({'i': vp_i, 'j': vp_j, 'z': z, 'p_raw': p_raw, 'p_bonf': p_bonf, 'sig_5pct': p_bonf < 0.05})
    print(f"{vp_i:<12} {vp_j:<12} {z:>+8.3f} {p_raw:>10.4f} {p_bonf:>10.4f} {sig:>8}")

results['dunn_bonferroni'] = dunn_results

# ===================================================================
# 3. SCORE COMPUESTO S = mu - k*sigma  (k=1)
# ===================================================================
print("\n" + "="*80)
print("3. SCORE COMPUESTO S = mu - k*sigma (k=1)")
print("="*80)

k = 1.0
score_vp = []
for vp in vps:
    rows = [r for r in vp_rows if r['VP']==vp]
    lrac_vals = np.array([r['LRAC4P'] for r in rows])
    mu = lrac_vals.mean()
    sigma = lrac_vals.std(ddof=1)
    S = mu - k*sigma
    score_vp.append({'vp': vp, 'mu': mu, 'sigma': sigma, 'S': S})

score_vp_sorted = sorted(score_vp, key=lambda x: x['S'], reverse=True)
print(f"\n{'VP':<12} {'µ_anual':>10} {'σ':>10} {'S=µ-σ':>10}  Ranking")
for i, s in enumerate(score_vp_sorted, 1):
    star = ' ⭐ MEJOR' if i==1 else ''
    print(f"{s['vp']:<12} {s['mu']*100:>9.2f}% {s['sigma']*100:>9.2f}pp {s['S']*100:>9.2f}%  #{i}{star}")

results['score_compuesto'] = [{'vp': s['vp'], 'mu': s['mu'], 'sigma': s['sigma'], 'S': s['S']} for s in score_vp_sorted]

print(f"\nInterpretacion: si premiamos estabilidad (k=1), OPERACIONES emerge como el sistema mas confiable.")
print(f"Aunque SUPPLY tiene µ ligeramente superior, su sigma lo penaliza fuerte.")

# ===================================================================
# 4. CONTROL CHART SHEWHART (I-Chart) sobre LRAC-4P corporativo
# ===================================================================
print("\n" + "="*80)
print("4. CONTROL CHART SHEWHART (I-Chart) - LRAC-4P corporativo")
print("="*80)

# Datos: 16 puntos (4 VPs × 4 meses)
all_points = [(r['VP'], r['MES'], r['MESNUM'], r['LRAC4P']) for r in vp_rows]
all_points.sort(key=lambda x: (x[2], x[0]))

vals_all = np.array([p[3] for p in all_points])
mu_global = vals_all.mean()
# Estimador de sigma robusto: mediana del rango movil entre puntos consecutivos
# Pero para I-Chart con n=16, uso sigma muestral
sigma_global = vals_all.std(ddof=1)

UCL = mu_global + 3*sigma_global
LCL = mu_global - 3*sigma_global
UWL = mu_global + 2*sigma_global  # warning
LWL = mu_global - 2*sigma_global

print(f"\nMedia global (centerline): µ = {mu_global*100:.2f}%")
print(f"Desv std global: σ = {sigma_global*100:.2f}pp")
print(f"UCL (µ+3σ) = {UCL*100:.2f}%")
print(f"LCL (µ-3σ) = {LCL*100:.2f}%")
print(f"UWL (µ+2σ) = {UWL*100:.2f}%")
print(f"LWL (µ-2σ) = {LWL*100:.2f}%")

# Puntos fuera de control
out_ucl = [(p[0], p[1], p[3]) for p in all_points if p[3] > UCL]
out_lcl = [(p[0], p[1], p[3]) for p in all_points if p[3] < LCL]
out_uwl = [(p[0], p[1], p[3]) for p in all_points if UWL < p[3] <= UCL]
out_lwl = [(p[0], p[1], p[3]) for p in all_points if LCL <= p[3] < LWL]

print(f"\nPuntos fuera de UCL: {len(out_ucl)}")
for vp, mes, v in out_ucl:
    print(f"  {vp:<12} {mes:<10} = {v*100:.2f}%")
print(f"Puntos fuera de LCL: {len(out_lcl)}")
for vp, mes, v in out_lcl:
    print(f"  {vp:<12} {mes:<10} = {v*100:.2f}%")
print(f"Puntos en zona de warning (entre 2σ y 3σ): UCL={len(out_uwl)}, LCL={len(out_lwl)}")
for vp, mes, v in out_uwl:
    print(f"  WARNING superior: {vp:<12} {mes:<10} = {v*100:.2f}%")
for vp, mes, v in out_lwl:
    print(f"  WARNING inferior: {vp:<12} {mes:<10} = {v*100:.2f}%")

results['shewhart'] = {
    'mu': mu_global, 'sigma': sigma_global,
    'UCL': UCL, 'LCL': LCL, 'UWL': UWL, 'LWL': LWL,
    'out_ucl': [(p[0], p[1], p[3]) for p in all_points if p[3] > UCL],
    'out_lcl': [(p[0], p[1], p[3]) for p in all_points if p[3] < LCL],
    'out_uwl': [(p[0], p[1], p[3]) for p in all_points if UWL < p[3] <= UCL],
    'out_lwl': [(p[0], p[1], p[3]) for p in all_points if LCL <= p[3] < LWL],
}

# ===================================================================
# 5. TENDENCIA MANN-KENDALL aproximada con Kendall's tau por VP
# ===================================================================
print("\n" + "="*80)
print("5. TENDENCIA MANN-KENDALL (Kendall's tau) - LRAC-4P por VP")
print("="*80)

# Mann-Kendall: estadistico S = sum(sgn(x_j - x_i)) para j>i
# Con n=4 puntos por VP, la potencia es muy baja, pero es ilustrativo
mk_results = []
print(f"\n{'VP':<12} {'tau':>8} {'p':>8} {'tendencia':>12}")
for vp in vps:
    rows = sorted([r for r in vp_rows if r['VP']==vp], key=lambda x: x['MESNUM'])
    vals = [r['LRAC4P'] for r in rows]
    x = list(range(1, len(vals)+1))
    tau, p_mk = stats.kendalltau(x, vals)
    direction = '↑ creciente' if tau > 0 else ('↓ decreciente' if tau < 0 else '→ plana')
    mk_results.append({'vp': vp, 'tau': tau, 'p': p_mk, 'direction': direction})
    print(f"{vp:<12} {tau:>+8.3f} {p_mk:>8.4f} {direction:>14}")

results['mann_kendall'] = mk_results
print(f"\nNota: con n=4 puntos la potencia estadistica es baja; se reporta como diagnostico.")

# ===================================================================
# 6. MATRIZ DE CORRELACION 7x7 ENTRE HERRAMIENTAS
# ===================================================================
print("\n" + "="*80)
print("6. MATRIZ DE CORRELACION 7x7 (Pearson) entre herramientas")
print("="*80)

indicadores = ['FTO', '3Q', 'CCV', 'ACC', 'NMAP', 'VEA', 'DESEMP']
data = np.array([[r[i] for i in indicadores] for r in ger_rows])
corr = np.corrcoef(data.T)

print(f"\n{'':>8}", end='')
for i in indicadores: print(f"{i:>8}", end='')
print()
for i, ind_i in enumerate(indicadores):
    print(f"{ind_i:>8}", end='')
    for j in range(len(indicadores)):
        print(f"{corr[i,j]:>+8.2f}", end='')
    print()

# Top 3 correlaciones positivas y negativas
pairs_c = []
for i in range(len(indicadores)):
    for j in range(i+1, len(indicadores)):
        pairs_c.append((indicadores[i], indicadores[j], corr[i,j]))
pairs_c.sort(key=lambda x: x[2], reverse=True)
print("\nTop 5 correlaciones positivas:")
for a,b,c in pairs_c[:5]:
    print(f"  {a:<8} -- {b:<8} = {c:+.3f}")
print("\nTop 5 correlaciones negativas:")
for a,b,c in pairs_c[-5:]:
    print(f"  {a:<8} -- {b:<8} = {c:+.3f}")

results['correlation_matrix'] = corr.tolist()
results['indicadores'] = indicadores

# ===================================================================
# GUARDAR RESULTADOS
# ===================================================================
with open(OUT_DIR + '/resultados_inferencial.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, indent=2, default=str, ensure_ascii=False)
print(f"\nResultados guardados en {OUT_DIR}/resultados_inferencial.json")


# ===================================================================
# FIGURAS NUEVAS
# ===================================================================

# FIG 11: Boxplot + resultado Kruskal-Wallis
fig, ax = plt.subplots(figsize=(7, 4.5))
data_plot = [samples_vp[vp] for vp in vps]
bp = ax.boxplot(data_plot, tick_labels=vps, patch_artist=True, widths=0.55,
                medianprops={'color':'black','linewidth':2})
for patch, vp in zip(bp['boxes'], vps):
    patch.set_facecolor(COLORES_VP[vp])
    patch.set_alpha(0.7)
ax.axhline(0.70, color='gray', linestyle='--', alpha=0.6, label='Meta 70%')
# Anotaciones
for i, vp in enumerate(vps):
    vals = samples_vp[vp]
    ax.text(i+1, max(vals)+0.005, f'n={len(vals)}', ha='center', fontsize=8)
# Cuadro de resultado test
sig_text = 'p<0.05 (significativo)' if p_kw < 0.05 else 'p≥0.05 (no significativo)'
ax.text(0.02, 0.97, f'Kruskal-Wallis\nH = {H:.3f}\np = {p_kw:.4f}\n{sig_text}',
        transform=ax.transAxes, va='top', fontsize=8,
        bbox=dict(boxstyle='round', facecolor='white', edgecolor='gray', alpha=0.9))
ax.set_ylabel('LRAC-4P por gerencia-mes')
ax.set_title('Figura 11. Comparacion no-parametrica entre VPs (Kruskal-Wallis)')
ax.grid(alpha=0.3, axis='y')
plt.savefig(FIG_DIR + '/fig11_kruskal_wallis.png')
plt.close()
print('Fig 11 OK')

# FIG 12: Control chart Shewhart
fig, ax = plt.subplots(figsize=(10, 4.5))
x = range(1, len(all_points)+1)
y = [p[3]*100 for p in all_points]
labels = [f"{p[0][:3]}\n{p[1][:3]}" for p in all_points]
colors_pts = [COLORES_VP[p[0]] for p in all_points]

# Bandas
ax.axhline(mu_global*100, color='black', linewidth=1.5, label=f'µ = {mu_global*100:.2f}%')
ax.axhline(UCL*100, color='red', linewidth=1.2, linestyle='--', label=f'UCL (µ+3σ)')
ax.axhline(LCL*100, color='red', linewidth=1.2, linestyle='--', label=f'LCL (µ-3σ)')
ax.fill_between(x, UWL*100, UCL*100, alpha=0.1, color='orange', label='Warning ±2σ')
ax.fill_between(x, LCL*100, LWL*100, alpha=0.1, color='orange')

for xi, yi, c, p in zip(x, y, colors_pts, all_points):
    ax.plot(xi, yi, 'o', color=c, markersize=8, markeredgecolor='black')
    if yi/100 > UWL or yi/100 < LWL:
        ax.annotate(f"{p[0][:4]}\n{p[1][:3]}", xy=(xi, yi),
                    xytext=(0, 12), textcoords='offset points',
                    fontsize=7, ha='center',
                    arrowprops=dict(arrowstyle='->', color='black', lw=0.5))

ax.set_xticks(x)
ax.set_xticklabels(labels, fontsize=6)
ax.set_ylabel('LRAC-4P (%)')
ax.set_title('Figura 12. I-Chart Shewhart corporativo - 16 puntos VP x mes (Ene-Abr 2026)')
ax.legend(loc='upper right', fontsize=7, ncol=2)
ax.set_ylim(55, 82)
ax.grid(alpha=0.3, axis='y')
plt.tight_layout()
plt.savefig(FIG_DIR + '/fig12_shewhart.png')
plt.close()
print('Fig 12 OK')

# FIG 13: Matriz correlacion 7x7
fig, ax = plt.subplots(figsize=(7, 6))
im = ax.imshow(corr, cmap='RdBu_r', vmin=-1, vmax=1, aspect='auto')
ax.set_xticks(range(len(indicadores)))
ax.set_xticklabels(indicadores, rotation=45, ha='right')
ax.set_yticks(range(len(indicadores)))
ax.set_yticklabels(indicadores)
for i in range(len(indicadores)):
    for j in range(len(indicadores)):
        color = 'white' if abs(corr[i,j]) > 0.5 else 'black'
        ax.text(j, i, f'{corr[i,j]:.2f}', ha='center', va='center', color=color, fontsize=9)
cbar = plt.colorbar(im, ax=ax, fraction=0.04)
cbar.set_label('Coef. Pearson')
ax.set_title('Figura 13. Matriz de correlacion entre las 7 herramientas LRAC (n=48 gerencia-mes)')
plt.savefig(FIG_DIR + '/fig13_correlacion.png')
plt.close()
print('Fig 13 OK')

# FIG 14: Score compuesto
fig, ax = plt.subplots(figsize=(8, 4))
xpos = np.arange(len(vps))
mus = [s['mu']*100 for s in score_vp_sorted]
sigmas = [s['sigma']*100 for s in score_vp_sorted]
Ss = [s['S']*100 for s in score_vp_sorted]
vps_sorted = [s['vp'] for s in score_vp_sorted]
colors = [COLORES_VP[vp] for vp in vps_sorted]

bars_mu = ax.bar(xpos - 0.25, mus, width=0.25, label='µ (promedio)', color=colors, alpha=0.85, edgecolor='black')
bars_s = ax.bar(xpos, Ss, width=0.25, label='S = µ − σ (score compuesto)', color=colors, alpha=0.4, edgecolor='black', hatch='///')
# Error bar sigma
ax.errorbar(xpos + 0.25, mus, yerr=sigmas, fmt='ko', label='σ (desv std)', capsize=4)

for i, (m, s) in enumerate(zip(mus, Ss)):
    ax.text(i-0.25, m+0.3, f'{m:.1f}%', ha='center', fontsize=8, fontweight='bold')
    ax.text(i, s+0.3, f'{s:.1f}%', ha='center', fontsize=8, fontweight='bold')

ax.axhline(70, color='gray', linestyle='--', alpha=0.6)
ax.set_xticks(xpos)
ax.set_xticklabels(vps_sorted)
ax.set_ylabel('% LRAC-4P')
ax.set_title('Figura 14. Score compuesto S = µ − σ premiando estabilidad (k=1)')
ax.legend(fontsize=8, loc='upper right')
ax.set_ylim(50, 75)
ax.grid(alpha=0.3, axis='y')
plt.savefig(FIG_DIR + '/fig14_score_compuesto.png')
plt.close()
print('Fig 14 OK')

print('\nTodas las figuras nuevas guardadas en', FIG_DIR)
print('\n=== ITER 2A COMPLETADA ===')
