"""
ITER 6C - Analisis Excel extra
Profundiza en los 3 indicadores no explorados por el reto pedido (CCV, VEA P1, DESEMPENO)
+ patrones cross-indicador a nivel gerencia.
"""
import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import numpy as np
from collections import defaultdict
from statistics import mean, stdev

EXCEL = r"C:/Users/Emanuel/Downloads/00 Data (10).xlsx"
OUT = r"C:/Users/Emanuel/.gemini/antigravity/scratch/n8n/lasbambas-lrac-2026/analysis"

wb = openpyxl.load_workbook(EXCEL, data_only=True)

base = []
for row in wb['BASE'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        base.append({'VP': row[0], 'GER': row[1], 'PILAR': row[2],
                     'HERR': row[3], 'VALOR': row[4], 'MES': row[5]})

ger_rows = []
for row in wb['LRACxGerencia'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        ger_rows.append({'VP': row[0], 'GER': row[1], 'MES': row[3],
                         '3Q': row[6], 'ACC': row[7], 'CCV': row[8],
                         'DESEMP': row[9], 'FTO': row[10], 'NMAP': row[11],
                         'VEA': row[12], 'LRAC4P': row[17]})

# ===================================================================
# Análisis 1: Peor gerencia en CCV, VEA P1, DESEMPEÑO (los 3 no
# explorados por el reto pedido)
# ===================================================================
print("="*80)
print("ANALISIS EXTRA 1 - PEORES GERENCIAS POR HERRAMIENTAS NO EXPLORADAS (ENERO)")
print("="*80)

for h in ['CCV', 'VEA P1', 'DESEMPEÑO']:
    enero_h = [r for r in base if r['MES']=='ENERO' and r['HERR']==h and r['VALOR'] is not None]
    enero_h.sort(key=lambda r: r['VALOR'])
    print(f"\n{h}: PEOR = {enero_h[0]['GER']} (VP {enero_h[0]['VP']}) con {enero_h[0]['VALOR']*100:.1f}%")
    print(f"  Media: {mean(r['VALOR'] for r in enero_h)*100:.2f}%")
    print(f"  Top 3 peores:")
    for r in enero_h[:3]:
        print(f"    {r['VP']:<12} | {r['GER']:<22} | {r['VALOR']*100:5.1f}%")

# ===================================================================
# Análisis 2: Cross-indicator patterns - gerencias que están en el bottom
# en multiples indicadores simultáneamente
# ===================================================================
print("\n" + "="*80)
print("ANALISIS EXTRA 2 - GERENCIAS CRITICAS CROSS-INDICADOR (ENERO)")
print("="*80)

herramientas = ['FTO', '3Q', 'CCV', 'ACC', 'NMAP', 'VEA P1', 'DESEMPEÑO']
# Para cada herramienta en enero, calcular ranking
enero = [r for r in base if r['MES']=='ENERO' and r['VALOR'] is not None]
rankings = defaultdict(list)
for h in herramientas:
    h_data = sorted([r for r in enero if r['HERR']==h], key=lambda r: r['VALOR'])
    for pos, r in enumerate(h_data, 1):
        rankings[(r['VP'], r['GER'])].append((h, pos, r['VALOR']))

# Sumar posiciones de ranking para cada gerencia (menor suma = mas debil overall)
ger_score = []
for (vp, ger), rks in rankings.items():
    avg_pos = mean(p for _, p, _ in rks)
    avg_val = mean(v for _, _, v in rks)
    in_bottom3 = sum(1 for _, p, _ in rks if p <= 3)
    ger_score.append({'vp': vp, 'ger': ger, 'avg_pos': avg_pos,
                       'avg_val': avg_val, 'bottom3_count': in_bottom3, 'rks': rks})

ger_score.sort(key=lambda x: x['avg_pos'])
print(f"\nGerencias con peor desempeno PROMEDIO en enero (ranking 1 = peor por herramienta):")
print(f"{'VP':<12} {'GERENCIA':<22} {'pos_media':>11} {'%_media':>9} {'bottom3':>9}")
for g in ger_score[:6]:
    print(f"{g['vp']:<12} {g['ger']:<22} {g['avg_pos']:>11.2f} {g['avg_val']*100:>8.2f}% {g['bottom3_count']:>9}")

print(f"\n>>> Gerencia mas critica overall enero: {ger_score[0]['ger']} (VP {ger_score[0]['vp']})")
print(f"    en bottom-3 de {ger_score[0]['bottom3_count']} de 7 herramientas")
print(f"    Detalle por herramienta (ranking 1=peor de 12):")
for h, p, v in sorted(ger_score[0]['rks']):
    flag = ' <-- bottom-3' if p <= 3 else ''
    print(f"      {h:<10} pos #{p:>2} ({v*100:5.1f}%){flag}")

# ===================================================================
# Análisis 3: Estabilidad de pilares por gerencia (CV intra-gerencia
# a lo largo de 4 meses)
# ===================================================================
print("\n" + "="*80)
print("ANALISIS EXTRA 3 - ESTABILIDAD INTRA-GERENCIA (CV % a lo largo del 2026)")
print("="*80)

ger_stab = defaultdict(list)
for r in ger_rows:
    ger_stab[(r['VP'], r['GER'])].append(r['LRAC4P'])

stab_list = []
for (vp, ger), vals in ger_stab.items():
    mu = mean(vals)
    sd = stdev(vals) if len(vals) > 1 else 0
    cv = sd / mu * 100 if mu else 0
    stab_list.append({'vp': vp, 'ger': ger, 'mu': mu, 'sd': sd, 'cv': cv})

stab_list.sort(key=lambda x: x['cv'])
print(f"\nGerencias MAS estables (CV % mas bajo):")
print(f"{'VP':<12} {'GERENCIA':<22} {'CV%':>8} {'sigma':>8}")
for g in stab_list[:5]:
    print(f"{g['vp']:<12} {g['ger']:<22} {g['cv']:>7.2f} {g['sd']*100:>7.2f}")

print(f"\nGerencias MAS volatiles (CV % mas alto):")
for g in stab_list[-5:]:
    print(f"{g['vp']:<12} {g['ger']:<22} {g['cv']:>7.2f} {g['sd']*100:>7.2f}")

# ===================================================================
# Análisis 4: Heatmap herramienta x VP enero
# ===================================================================
print("\n" + "="*80)
print("ANALISIS EXTRA 4 - HEATMAP HERRAMIENTA x VP - ENERO 2026")
print("="*80)

vps = sorted(set(r['VP'] for r in base))
print(f"\n{'':>12}", end='')
for vp in vps:
    print(f"{vp:>14}", end='')
print()
for h in herramientas:
    print(f"{h:<12}", end='')
    for vp in vps:
        vals = [r['VALOR'] for r in enero if r['HERR']==h and r['VP']==vp]
        if vals:
            mu = mean(vals) * 100
            print(f"{mu:>13.1f}%", end='')
        else:
            print(f"{'-':>14}", end='')
    print()

# Identificar el peor par (herramienta, VP)
peor_par = None
peor_v = 1.0
for h in herramientas:
    for vp in vps:
        vals = [r['VALOR'] for r in enero if r['HERR']==h and r['VP']==vp]
        if vals:
            mu = mean(vals)
            if mu < peor_v:
                peor_v = mu
                peor_par = (h, vp)
print(f"\n>>> Combinacion (herramienta, VP) mas debil en enero: {peor_par} con {peor_v*100:.2f}%")

# ===================================================================
# GUARDAR JSON
# ===================================================================
out = {
    'cross_indicator_critical': [
        {'vp': g['vp'], 'gerencia': g['ger'], 'pos_media': g['avg_pos'],
         'valor_medio': g['avg_val'], 'bottom3_count': g['bottom3_count']}
        for g in ger_score[:6]
    ],
    'estabilidad_intra_gerencia': stab_list,
    'peor_par_herramienta_vp_enero': {'herramienta': peor_par[0], 'vp': peor_par[1], 'valor': peor_v},
}
with open(OUT + '/resultados_extra.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, indent=2, default=str, ensure_ascii=False)

print(f"\nResultados guardados en {OUT}/resultados_extra.json")
print("\n=== ITER 6C COMPLETADA ===")
